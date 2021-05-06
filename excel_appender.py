import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, colors
from openpyxl.cell.cell import Cell

methods = {
    "_comment": "This json file sets ranking colors for translation methods in excel spread sheets.",
    "EQUAL": "00FF00",
    "EQUAL_LOWERCASE": "008000",
    "EQUAL_STRIPPED": "008080",
    "EQUAL_KEYS": "FF00FF",
    "SIMILAR_LCS": "080000"
}

class Ranked_Translations:
    def __init__(self, rt_filename : str):
        self.workbook = Workbook()
        self.ws = self.workbook.active
        self.filename = rt_filename.split(".")[0]
        with open(rt_filename, "r", encoding="utf-8") as js:
            dc = json.load(js)
        self.rts = dc
    
    def rank_cells(self, rank, quality):
        c = Cell(self.ws, column="A", row=1, value=rank)
        color = methods[rank]

        if quality == 0:
            # If only 1 locale consists of translation, make color "Reddish".
            color = "FF" + color[2:]
        elif quality > 1:
            # If translation is consistent across 2 or more locales, then GREEN.
            color = "00FF00"

        # Fill background of cell with color
        pf = PatternFill("solid", fgColor=colors.Color(color))
        c.fill = pf
        return c

    def to_excel(self):
        for row in self.rts:
            self.ws.append( [row, self.rts[row]["value"], self.rank_cells(self.rts[row]["rank"], self.rts[row]["quality"])] )
    
        self.workbook.save(filename=f"{self.filename}.xlsx")
'''
rt = Ranked_Translations("test.json")
rt.to_excel()
'''