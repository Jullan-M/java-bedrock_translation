import json
from copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill, colors

from utilities import dict2json, dict2lang, lang2dict

methods = {
    "_comment": "This json file sets ranking colors for translation methods in excel spread sheets.",
    "EQUAL": "00FF00",
    "EQUAL_LOWERCASE": "008000",
    "EQUAL_STRIPPED": "008080",
    "EQUAL_KEYS": "FF00FF",
    "SIMILAR_LCS": "080000",
}


class Ranked_Translations:
    def __init__(self, rts: dict):
        self.workbook = Workbook()
        self.ws = self.workbook.active
        self.rts = rts  # Dictionary

    def rank_cells(self, rank, quality):
        c = Cell(self.ws, column="A", row=1, value=rank)
        color = methods[rank]
        if quality == 0:
            # If only 1 locale consists of translation, make color "Reddish".
            color = "FF" + color[2:]
        elif rank != "EQUAL_LOWERCASE" and quality >= 1:
            # If translation is consistent across 2 or more locales, then GREEN.
            color = "00FF00"
        elif quality > 1:
            color = "00FF00"

        # Fill background of cell with color
        pf = PatternFill("solid", fgColor=colors.Color(color))
        c.fill = pf
        return c

    def to_excel(self, filename):
        for row in self.rts:
            self.ws.append(
                [
                    row,
                    self.rts[row]["value"],
                    self.rank_cells(self.rts[row]["rank"], self.rts[row]["quality"]),
                ]
            )

        self.workbook.save(filename=filename)


def excel2dict(filename: str):
    # Convert   key.goes.here   value     formatted spreadsheets to dict
    wb = load_workbook(filename)  # Workbook
    sh = wb.active  # Sheet
    dc = {}
    for c1, c2 in zip(sh["A"], sh["B"]):
        k, v = c1.value, c2.value
        if k and not (v == None):
            dc[k] = v
    return dc


def dict2excel(dictionary: str, out_file: str, template: str = ""):
    wb = Workbook()
    sh = wb.active
    if template:
        with open(template, "r", encoding="utf-8") as tp:
            for line in tp:
                strip = line.strip()
                if strip and strip[0] != "#" and strip[1] != "#":
                    k = line.split("=", maxsplit=1)[0]
                    if k in dictionary:
                        sh.append([k, dictionary[k]])
                    else:
                        sh.append([k, ""])
                else:
                    sh.append([])
    else:
        for k, v in dictionary.items():
            sh.append([k, v])
    wb.save(filename=out_file)


def comments2excel(outfile: str, template: str):
    wb = Workbook()
    sh = wb.active
    with open(template, "r", encoding="utf-8") as tp:
        for line in tp:
            strip = line.strip()
            if strip and (strip[0] == "#" or strip[1] == "#"):
                k = line.split("=", maxsplit=1)[0]
                sh.append([k, ""])
            else:
                sh.append([])


def excel2lang(filename: str, out_file: str, template: str = ""):
    dc = excel2dict(filename)
    dict2lang(dc, out_file, template)


def excel2json(filename: str, out_file: str):
    dc = excel2dict(filename)
    dict2json(dc, out_file)


def formatize_excel(filename: str, out_file: str, template: str):
    wb1 = load_workbook(filename)
    sh1 = wb1.active
    cell_dc = {}
    for c1, c2 in zip(sh1["A"], sh1["B"]):
        k, v = c1.value, c2.value
        if k and not (v == None):
            cell_dc[k] = c2

    wb2 = Workbook()
    sh2 = wb2.active
    with open(template, "r", encoding="utf-8") as tp:
        for line in tp:
            strip = line.strip()
            if strip and strip[0] != "#" and strip[1] != "#":
                k = line.split("=", maxsplit=1)[0]
                sh2.append([k, ""])
            else:
                sh2.append([])

    for c1, c2 in zip(sh2["A"], sh2["B"]):
        k = c1.value
        if k and k in cell_dc:
            cell = cell_dc[k]
            c2.value = cell.value
            c2.fill = copy(cell.fill)

    wb2.save(filename=out_file)
